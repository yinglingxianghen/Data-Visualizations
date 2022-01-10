import json
import requests
import openpyxl

def get_data():
    headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36",
            'Referer':'https://map.baidu.com/@12949550.923158279,3712445.9716704674,6.28z',
            "Cookie":"你的cookie"
        }
    url = "https://map.baidu.com/?newmap=1&reqflag=pcmap&biz=1&from=webmap&da_par=direct&pcevaname=pc4.1&qt=s&da_src=searchBox.button&wd=%E7%81%AB%E9%94%85%E5%BA%97&c=1&src=0&wd2=&pn=0&sug=0&l=6&b=(10637065.476146251,2368134.592189369;12772445.910805061,5056757.351151566)&from=webmap&biz_forward={%22scaler%22:1,%22styles%22:%22pl%22}&sug_forward=&auth=NTSwAZUMzIaTTdWD4WAv0731cWF3MQEauxLxREHzERRtykiOxAXXw1GgvPUDZYOYIZuVt1cv3uVtGccZcuVtPWv3GuztQZ3wWvUvhgMZSguxzBEHLNRTVtcEWe1GD8zv7ucvY1SGpuxVthgW1aDeuxtf0wd0vyMySFIAFM7ueh33uTtAffbDF&seckey=c6d9c7e05d7e627c56ed46fab5d7c5c792064779599d5e12b955a6f18a1204375d1588206c94d22e4bdd1ade0ad06e78c21917e24c6223b96bc51b75ca38651a1b203a0609f126163c5e82fd0549a068e537303424837ab798acfc9088e5d76a66451c20ebd9599b41c9b4f1371850d20fa442ad464712f54c912422f4fa20b3052f8bb810f30d41c7c0e55af68f9d9d973537f03d0aa0a1d1617d78cae29b49c64c2d2dc3f44cf0f8799234b124a7a2dec18bfa011e097e31a508eae37b8603f97df8f935f04b3652f190eac52d04816f302a582c53971e515ff2e0e2b4cc30446e0bee48d51c4be8b6fe4185589ed9&device_ratio=1&tn=B_NORMAL_MAP&nn=0&u_loc=12677548,2604239&ie=utf-8&t=1618452491622"

    response = requests.get(url,headers=headers).json()
    response = response['more_city']
    outwb_p = openpyxl.Workbook()
    outws_p = outwb_p.create_sheet(index=0)
    outws_p.cell(row=1, column=1, value="省份")
    outws_p.cell(row=1, column=2, value="数量")

    outwb_c = openpyxl.Workbook()
    outws_c = outwb_c.create_sheet(index=0)
    outws_c.cell(row=1, column=1, value="城市")
    outws_c.cell(row=1, column=2, value="数量")

    count_p=2
    count_c=2
    for i in response:
        city = i['city']
        print(i['province'])
        print(i['num'])
        ###保存全国省份火锅数量-李运辰
        outws_p.cell(row=count_p, column=1, value=str(i['province']))
        outws_p.cell(row=count_p, column=2, value=str(i['num']))
        count_p = count_p+1
        for j in city:
            print(j['name'])
            print(j['num'])
            ###保存全国城市火锅数量-李运辰.xls
            outws_c.cell(row=count_c, column=1, value=str(j['name']))
            outws_c.cell(row=count_c, column=2, value=str(j['num']))
            count_c = count_c+1

    outwb_p.save("全国省份火锅数量-李运辰.xls")  # 保存
    outwb_c.save("全国城市火锅数量-李运辰.xls")  # 保存

import pandas as pd

def drawmap1():

    datafile = u'全国省份火锅数量-李运辰.xls'
    data = pd.read_excel(datafile)
    attr = data['省份'].tolist()
    value = data['数量'].tolist()
    name = []
    for i in attr:
        if "省" in i:
            name.append(i.replace("省",""))
        else:
            name.append(i)
    from pyecharts import options as opts
    from pyecharts.charts import Map
    from pyecharts.faker import Faker
    c = (
        Map()
            .add("数量", [list(z) for z in zip(name, value)], "china")
            .set_global_opts(title_opts=opts.TitleOpts(title="全国火锅店数量分布情况"))
            .render("全国火锅店数量分布情况.html")
    )
def drawmap2():
    from pyecharts import options as opts
    from pyecharts.charts import Map
    from pyecharts.faker import Faker
    import os

    datafile = u'全国城市火锅数量-李运辰.xls'
    data = pd.read_excel(datafile)
    city = data['城市'].tolist()
    values2 = data['数量'].tolist()

    ###四川
    name = []
    value = []
    flag = 0
    for i in range(0,len(city)):
        if city[i] =="绵阳市":
            flag = 1
        if flag:
            name.append(city[i])
            value.append(int(values2[i]))

        if city[i] =="甘孜藏族自治州":
            name.append(city[i])
            value.append(int(values2[i]))
            break
    c = (
        Map()
            .add("四川火锅店数量分布", [list(z) for z in zip(name, value)], "四川")
            .set_global_opts(
            title_opts=opts.TitleOpts(title="四川火锅店数量分布"), visualmap_opts=opts.VisualMapOpts()
        )
            .render("四川火锅店数量分布.html")
    )

def drawmap3():
    from pyecharts import options as opts
    from pyecharts.charts import Map
    from pyecharts.faker import Faker
    import os

    datafile = u'全国城市火锅数量-李运辰.xls'
    data = pd.read_excel(datafile)
    city = data['城市'].tolist()
    values2 = data['数量'].tolist()

    ###广东
    name = []
    value = []
    flag = 0
    for i in range(0,len(city)):
        if city[i] =="广州市":
            flag = 1
        if flag:
            name.append(city[i])
            value.append(int(values2[i]))

        if city[i] =="云浮市":
            name.append(city[i])
            value.append(int(values2[i]))
            break
    c = (
        Map()
            .add("广东火锅店数量分布", [list(z) for z in zip(name, value)], "广东")
            .set_global_opts(
            title_opts=opts.TitleOpts(title="广东火锅店数量分布"), visualmap_opts=opts.VisualMapOpts()
        )
            .render("广东火锅店数量分布.html")
    )

from pyecharts.charts import Geo
from pyecharts import options as opts
from pyecharts.globals import ChartType

def d3():

    datafile = u'全国省份火锅数量-李运辰.xls'
    df = pd.read_excel(datafile)
    province_distribution = df[['省份', '数量']].values.tolist()
    geo = Geo()
    geo.set_series_opts(label_opts=opts.LabelOpts(is_show=False))
    geo.add_schema(maptype="china")
    geo.set_global_opts(visualmap_opts=opts.VisualMapOpts(max_=110000))
    # 加入数据
    geo.add('全国火锅店数量分布情况图2', province_distribution, type_=ChartType.EFFECT_SCATTER)
    geo.render("全国火锅店数量分布情况图2.html")

#d3()
#drawmap2()
#drawmap3()
#drawmap1()